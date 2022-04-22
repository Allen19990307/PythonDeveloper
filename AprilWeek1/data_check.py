import openpyxl
import datetime
from openpyxl import Workbook
import pandas as pd
if __name__ == '__main__':
    path = 'D:\DeveloperLife\PythonDeveloper\AprilWeek1\idp调度执行情况V2022-04-06.xlsx'
    wb = openpyxl.load_workbook(path)
    print(type(wb))
    # for x in wb.sheetnames:
        # 把元数据的内容写入Excel模板里面
        # df = pd.read_excel(path, sheet_name=x, usecols=[0], header=1)
        # df.to_excel(x + '.xlsx', startrow=1, index=False)
        # 字段的设置
    wb1 = openpyxl.load_workbook(path)
    date_current = datetime.datetime.now().strftime('%Y%m%d')
    wb1.create_sheet(date_current)

    sheet1 = wb1.get_sheet_by_name('20220406')
    sheet2 = wb1.get_sheet_by_name(date_current)
    """将sheet1上的内容复制到sheet2上面  """
    for i, row in enumerate(sheet1.iter_rows()):
        for j, cell in enumerate(row):
            sheet2.cell(row=i + 1, column=j + 1, value=cell.value)
    # sheet = wb1.active
    # sheet['A1'] = '字段名'
    # sheet['B1'] = '类型'
    # sheet['B2'] = 'string'
    # sheet['C1'] = '精度1'
    # sheet['D1'] = '精度2'
    # sheet['E1'] = '字段描述'
    wb1.save(path)
    wb1.close()
    wb.close()



