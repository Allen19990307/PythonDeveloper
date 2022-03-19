import openpyxl
from openpyxl import Workbook
import pandas as pd
if __name__ == '__main__':
    path = 'D:\DeveloperLife\PythonDeveloper\MarchWeek2Office\dy_data\抖音表结构.xlsx'
    wb = openpyxl.load_workbook(path)
    print(type(wb))
    for x in wb.sheetnames:
        # 把元数据的内容写入Excel模板里面
        df = pd.read_excel(path, sheet_name=x, usecols=[0], header=1)
        df.to_excel(x + '.xlsx', startrow=1, index=False)
        # 字段的设置
        wb1 = openpyxl.load_workbook(x + '.xlsx')
        sheet = wb1.active
        sheet.title = '表结构'
        sheet['A1'] = '字段名'
        sheet['B1'] = '类型'
        sheet['B2'] = 'string'
        sheet['C1'] = '精度1'
        sheet['D1'] = '精度2'
        sheet['E1'] = '字段描述'
        wb1.save(x+'.xlsx')
        wb1.close()
    wb.close()



