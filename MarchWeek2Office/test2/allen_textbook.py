import openpyxl
from openpyxl import Workbook
import pandas as pd

if __name__ == '__main__':
    """加载allen.xlsx表格"""
    wb = openpyxl.load_workbook('allen.xlsx')
    """显示表的类型"""
    print(type(wb))
    """遍历Excel中的所有sheet名称"""
    for x in wb.sheetnames:
        """先读取Excel表格的，sheet_name是sheet页名字，usecols是选用第几列的元素，header是从第几个开始"""
        df = pd.read_excel('allen.xlsx', sheet_name=x, usecols=[0], header=1)
        """写入Excel表格，startrow表示从第几行开始，index使用False表示我们不需要前面一列的索引（前期会使用，后期再深入理解，慢慢来就好）"""
        df.to_excel(x + '.xlsx', startrow=1, index=False)
        """下面给创建的Excel文件赋值：举个栗子,A1表示为A列，第一行的元素"""
        wb1 = openpyxl.load_workbook(x + '.xlsx')
        sheet = wb1.active
        sheet.title = 'find your passion'
        sheet['A1'] = 'How'
        sheet['B1'] = 'bad'
        sheet['C1'] = 'do'
        sheet['D1'] = 'you'
        sheet['E1'] = 'want'
        sheet['F1'] = 'it'
        sheet['G1'] = '?'
        """赋值之后，我们需要保存该文件"""
        wb1.save(x+'.xlsx')
        """加载使用完后就关闭"""
        wb1.close()
    wb.close()