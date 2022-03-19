from openpyxl import load_workbook
wb=load_workbook('2.xlsx')
s2 = [
'inner_world',
'money_maker',
'billionaire_world',
]
for i in range(0,len(s2)):
    wb.create_sheet(index=i,title= s2[i])
wb.save('2.xlsx')