import openpyxl
from openpyxl import Workbook
import pandas as pd
if __name__ == '__main__':
    wb = openpyxl.load_workbook('../test1/mongodb数据字典(1).xlsx')
    print(type(wb))
    for x in wb.sheetnames:
        df = pd.read_excel('../mongodb数据字典(1).xlsx', sheet_name=x, usecols=[0], header=1)
        df.to_excel(x + '.xlsx', startrow=1, index=False)
        wb1 = openpyxl.load_workbook(x + '.xlsx')
        sheet = wb1.active
        sheet.title = '表结构'
        sheet['A1'] = '字段名'
        sheet['B1'] = '类型'
        sheet['B2'] = 'string'
        sheet['C1'] = '精度1'
        sheet['D1'] = '精度2'
        sheet['E1'] = '字段描述'
        wb1.save(x+'.xlsx')
        wb1.close()
    wb.close()

